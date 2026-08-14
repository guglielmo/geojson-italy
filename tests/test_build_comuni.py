import json

import openpyxl
import pytest

from scripts.build_comuni import (
    LEGACY_FIELDS,
    build_properties,
    index_layers,
    merge_legacy,
    read_catasto_codes,
    read_legacy_by_catasto,
)


@pytest.fixture
def elenco(tmp_path):
    """A minimal Elenco-comuni-italiani.xlsx with only the columns we read."""
    wb = openpyxl.Workbook()
    ws = wb.active
    header = [""] * 21
    header[4] = "Codice Comune formato alfanumerico"
    header[6] = "Denominazione in italiano"
    header[20] = "Codice Catastale del Comune"
    ws.append(header)
    for istat, name, catasto in [
        ("001001", "Agliè", "A074"),
        ("113001", "Aggius", "A069"),
    ]:
        row = [None] * 21
        row[4], row[6], row[20] = istat, name, catasto
        ws.append(row)
    path = tmp_path / "elenco.xlsx"
    wb.save(path)
    return path


def test_read_catasto_codes_maps_istat_to_catasto(elenco):
    assert read_catasto_codes(elenco) == {"001001": "A074", "113001": "A069"}


def test_read_catasto_codes_survives_a_moved_column(tmp_path):
    """ISTAT reorders this spreadsheet between editions.

    The columns are located by heading, so a move must be absorbed rather than
    silently producing a wrong join.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Codice Catastale del Comune", "Denominazione in italiano",
               "Codice Comune formato alfanumerico"])
    ws.append(["A074", "Agliè", "001001"])
    path = tmp_path / "moved.xlsx"
    wb.save(path)

    assert read_catasto_codes(path) == {"001001": "A074"}


def test_read_catasto_codes_fails_loudly_on_a_missing_column(tmp_path):
    """A renamed or dropped column must stop the build, not empty the join."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Denominazione in italiano", "Codice Comune formato alfanumerico"])
    ws.append(["Agliè", "001001"])
    path = tmp_path / "incomplete.xlsx"
    wb.save(path)

    with pytest.raises(ValueError, match="Codice Catastale del Comune"):
        read_catasto_codes(path)


def test_build_properties_carries_territorial_codes():
    comune = {
        "COMUNE": "Aggius",
        "PRO_COM_T": "113001",
        "COD_PROV": 113,
        "COD_REG": 20,
    }
    uts = {113: {"DEN_UTS": "Gallura Nord-Est Sardegna", "SIGLA": "OT"}}
    regions = {20: "Sardegna"}
    props = build_properties(comune, uts, regions, catasto="A069")

    assert props["name"] == "Aggius"
    assert props["com_istat_code"] == "113001"
    assert props["com_istat_code_num"] == 113001
    assert props["com_catasto_code"] == "A069"
    assert props["prov_istat_code"] == "113"
    assert props["prov_istat_code_num"] == 113
    assert props["prov_name"] == "Gallura Nord-Est Sardegna"
    assert props["prov_acr"] == "OT"
    assert props["reg_istat_code"] == "20"
    assert props["reg_istat_code_num"] == 20
    assert props["reg_name"] == "Sardegna"


def test_build_properties_zero_pads_codes():
    comune = {"COMUNE": "Agliè", "PRO_COM_T": "001001", "COD_PROV": 1, "COD_REG": 1}
    uts = {1: {"DEN_UTS": "Torino", "SIGLA": "TO"}}
    props = build_properties(comune, uts, {1: "Piemonte"}, catasto="A074")

    assert props["prov_istat_code"] == "001"
    assert props["reg_istat_code"] == "01"


def test_build_properties_preserves_the_published_key_order():
    """The published file has carried this key order since 2019.

    JSON objects are unordered by specification, but the file is diffed and read
    by people. Reordering would make the release diff unreadable for no gain,
    and the legacy fields are placed here as None so that merge_legacy can fill
    them in without moving them.
    """
    comune = {"COMUNE": "Agliè", "PRO_COM_T": "001001", "COD_PROV": 1, "COD_REG": 1}
    props = build_properties(comune, {1: {"DEN_UTS": "Torino", "SIGLA": "TO"}},
                             {1: "Piemonte"}, catasto="A074")

    assert list(props) == [
        "name", "op_id", "minint_elettorale", "minint_finloc",
        "prov_name", "prov_istat_code", "prov_istat_code_num", "prov_acr",
        "reg_name", "reg_istat_code", "reg_istat_code_num",
        "opdm_id", "com_catasto_code", "com_istat_code", "com_istat_code_num",
    ]


def test_read_legacy_by_catasto_keys_on_cadastral_code(tmp_path):
    previous = {
        "type": "FeatureCollection",
        "features": [
            {
                "type": "Feature",
                "properties": {
                    "name": "Aggius",
                    "com_catasto_code": "A069",
                    "com_istat_code": "090001",
                    "op_id": "1234",
                    "opdm_id": "abcd",
                    "minint_elettorale": "1010810010",
                    "minint_finloc": "1010810010",
                },
                "geometry": None,
            }
        ],
    }
    path = tmp_path / "previous.geojson"
    path.write_text(json.dumps(previous))

    legacy = read_legacy_by_catasto(path)

    # Keyed on the cadastral code, so the Sardinian recoding 090001 -> 113001
    # does not break the lookup.
    assert legacy["A069"]["op_id"] == "1234"
    assert legacy["A069"]["minint_finloc"] == "1010810010"
    assert "090001" not in legacy


def test_merge_legacy_fills_known_municipality():
    props = {"name": "Aggius", "com_catasto_code": "A069"}
    merged, missing = merge_legacy(props, {"A069": {f: "x" for f in LEGACY_FIELDS}})
    assert merged["op_id"] == "x"
    assert missing == []


def test_merge_legacy_leaves_new_municipality_null():
    props = {"name": "Bardello con Malgesso e Bregano", "com_catasto_code": "M441"}
    merged, missing = merge_legacy(props, {})
    assert all(merged[f] is None for f in LEGACY_FIELDS)
    assert missing == ["Bardello con Malgesso e Bregano"]


def test_index_layers_extracts_uts_and_regions():
    prov = {
        "features": [
            {"properties": {"COD_PROV": 113, "DEN_UTS": "Gallura Nord-Est Sardegna",
                            "SIGLA": "OT", "TIPO_UTS": "Provincia"}}
        ]
    }
    reg = {"features": [{"properties": {"COD_REG": 20, "DEN_REG": "Sardegna"}}]}
    uts, regions = index_layers(prov, reg)
    assert uts[113]["SIGLA"] == "OT"
    assert regions[20] == "Sardegna"
