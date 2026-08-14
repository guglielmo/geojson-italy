"""Build comuni.geojson from the ISTAT sources plus the previous release.

The join key between releases is the cadastral (Belfiore) code, never the ISTAT
code: the Sardinian reform of 1 January 2026 changed all 377 Sardinian ISTAT
codes with no overlap, and municipality names both collide and change.
"""

import json
from pathlib import Path

import openpyxl

# Headings in Elenco-comuni-italiani.xlsx. The columns are located by heading
# rather than by position: ISTAT reorders this spreadsheet between editions, and
# a hardcoded index would keep reading, silently joining the wrong column.
_HEADER_ISTAT = "Codice Comune formato alfanumerico"
_HEADER_CATASTO = "Codice Catastale del Comune"


def _resolve_columns(header):
    """Return (istat index, catasto index), located by column heading."""
    positions = {}
    for index, cell in enumerate(header):
        if cell is None:
            continue
        # Headings carry stray newlines and double spaces between editions.
        label = " ".join(str(cell).split())
        if label in (_HEADER_ISTAT, _HEADER_CATASTO):
            positions[label] = index
    missing = [h for h in (_HEADER_ISTAT, _HEADER_CATASTO) if h not in positions]
    if missing:
        raise ValueError(f"column headings not found in the spreadsheet: {missing}")
    return positions[_HEADER_ISTAT], positions[_HEADER_CATASTO]


def read_catasto_codes(path):
    """Return {istat_code: catasto_code} from the ISTAT spreadsheet."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    rows = wb[wb.sheetnames[0]].iter_rows(values_only=True)
    col_istat, col_catasto = _resolve_columns(next(rows))
    out = {}
    for row in rows:
        istat = row[col_istat]
        catasto = row[col_catasto]
        if istat and catasto:
            out[str(istat).strip()] = str(catasto).strip()
    return out


def build_properties(comune, uts_by_code, region_names, catasto):
    """Assemble the published property set for one municipality.

    Codes are carried both zero-padded as text and as integers, matching the
    schema this repository has always published. The key order reproduces the
    published file's, legacy fields included as None placeholders so that
    merge_legacy fills them in place rather than appending them.
    """
    prov_code = int(comune["COD_PROV"])
    reg_code = int(comune["COD_REG"])
    istat = str(comune["PRO_COM_T"]).strip()
    uts = uts_by_code[prov_code]

    return {
        "name": comune["COMUNE"],
        "op_id": None,
        "minint_elettorale": None,
        "minint_finloc": None,
        "prov_name": uts["DEN_UTS"],
        "prov_istat_code": f"{prov_code:03d}",
        "prov_istat_code_num": prov_code,
        "prov_acr": uts["SIGLA"],
        "reg_name": region_names[reg_code],
        "reg_istat_code": f"{reg_code:02d}",
        "reg_istat_code_num": reg_code,
        "opdm_id": None,
        "com_catasto_code": catasto,
        "com_istat_code": istat,
        "com_istat_code_num": int(istat),
    }


# Identifiers that exist only in previous releases of this repository. They are
# not published by ISTAT and cannot be derived, so they are carried across.
LEGACY_FIELDS = ("op_id", "opdm_id", "minint_elettorale", "minint_finloc")


def read_legacy_by_catasto(path):
    """Return {catasto_code: {legacy field: value}} from a previous release."""
    features = json.loads(Path(path).read_text())["features"]
    out = {}
    for feature in features:
        props = feature["properties"]
        catasto = props.get("com_catasto_code")
        if catasto:
            out[catasto] = {field: props.get(field) for field in LEGACY_FIELDS}
    return out


def merge_legacy(props, legacy_by_catasto):
    """Add the legacy identifiers to props.

    Returns (props, missing) where missing lists the names of municipalities
    with no previous entry. A genuinely new municipality has no openpolis or
    interior-ministry identifier yet; None is the honest value.

    Assigning to keys build_properties already placed leaves them where they
    are, which is what keeps the published key order intact.
    """
    found = legacy_by_catasto.get(props["com_catasto_code"])
    merged = dict(props)
    for field in LEGACY_FIELDS:
        merged[field] = found[field] if found else None
    return merged, ([] if found else [props["name"]])


def index_layers(prov_layer, reg_layer):
    """Index the province/UTS and region layers by their numeric codes."""
    uts = {
        int(f["properties"]["COD_PROV"]): f["properties"]
        for f in prov_layer["features"]
    }
    regions = {
        int(f["properties"]["COD_REG"]): f["properties"]["DEN_REG"]
        for f in reg_layer["features"]
    }
    return uts, regions


# Municipalities present in a boundary edition but absent from a later edition
# of the spreadsheet, because they were suppressed after the reference date. The
# cadastral code is taken from the previous release of this repository, where all
# three are still present.
CATASTO_OVERRIDES = {
    "018082": "E608",  # Lirio, absorbed by an existing municipality after 2026-01-01
    "024027": "C056",  # Castegnero, merged into Castegnero Nanto (024129)
    "024071": "F838",  # Nanto, same merger
}


def build(year, previous_path, out_path):
    """Build comuni.geojson for a reference year. Returns a report dict."""
    src = Path("build/istat") / str(year)
    comuni = json.loads((src / "Com.geojson").read_text())
    uts, regions = index_layers(
        json.loads((src / "ProvCM.geojson").read_text()),
        json.loads((src / "Reg.geojson").read_text()),
    )
    catasto_by_istat = read_catasto_codes(src / "elenco.xlsx")
    legacy = read_legacy_by_catasto(previous_path)

    features = []
    missing_catasto = []
    missing_legacy = []
    for feature in comuni["features"]:
        props = feature["properties"]
        istat = str(props["PRO_COM_T"]).strip()
        catasto = catasto_by_istat.get(istat) or CATASTO_OVERRIDES.get(istat)
        if catasto is None:
            # Present in the boundary edition but not in the later spreadsheet,
            # e.g. a municipality merged away after the reference date.
            missing_catasto.append((istat, props["COMUNE"]))
            catasto = None
        built = build_properties(props, uts, regions, catasto)
        built, missing = merge_legacy(built, legacy)
        missing_legacy.extend(missing)
        features.append(
            {"type": "Feature", "properties": built, "geometry": feature["geometry"]}
        )

    out = {
        "type": "FeatureCollection",
        "crs": {"type": "name", "properties": {"name": "EPSG:4326"}},
        "features": features,
    }
    Path(out_path).write_text(json.dumps(out, ensure_ascii=True))
    return {
        "count": len(features),
        "missing_catasto": missing_catasto,
        "missing_legacy": sorted(missing_legacy),
    }


if __name__ == "__main__":
    import sys

    report = build(sys.argv[1], "comuni.geojson.prev", "comuni.geojson")
    print(f"features: {report['count']}")
    print(f"without cadastral code: {len(report['missing_catasto'])}")
    for istat, name in report["missing_catasto"]:
        print(f"  {istat} {name}")
    print(f"without legacy identifiers: {len(report['missing_legacy'])}")
    for name in report["missing_legacy"]:
        print(f"  {name}")
